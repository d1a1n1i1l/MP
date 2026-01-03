import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse

def export_diploma_to_excel(diploma_disciplines, user, gpa):
    """Экспорт диплома в Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Диплом"
    
    # Заголовок
    ws.merge_cells('A1:D1')
    ws['A1'] = "АКАДЕМИЧЕСКИЙ ДИПЛОМ"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:D2')
    ws['A2'] = f"Студент: {user.get_full_name() or user.username}"
    ws['A2'].font = Font(size=12)
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Пустая строка
    ws.append([])
    
    # Заголовки таблицы
    headers = ["Семестр", "Дисциплина", "Вид контроля", "Оценка"]
    ws.append(headers)
    for cell in ws[4]:  # 4-я строка (заголовки)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Данные
    for d in diploma_disciplines:
        ws.append([
            d.semester.number,
            d.title,
            d.get_assessment_type_display(),
            d.actual_grade or "—"
        ])
    
    # Итог: средний балл
    ws.append([])
    ws.append(["Средний балл диплома", "", "", gpa])
    ws[f'A{ws.max_row}'].font = Font(bold=True)
    ws[f'D{ws.max_row}'].font = Font(bold=True, size=14, color="0000FF")
    
    # Автоширина колонок (исправление ошибки MergedCell)
    for col_idx, col in enumerate(ws.iter_cols(), start=1):
        max_length = 0
        column = openpyxl.utils.get_column_letter(col_idx)
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width if adjusted_width < 50 else 50
    
    return wb

def export_transcript_to_excel(disciplines, semester, user, gpa):
    """Экспорт ведомости в Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Семестр {semester.number}"
    
    # Заголовок
    ws.merge_cells('A1:E1')
    ws['A1'] = f"АКАДЕМИЧЕСКАЯ ВЕДОМОСТЬ - СЕМЕСТР {semester.number} ({semester.academic_year})"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:E2')
    ws['A2'] = f"Студент: {user.get_full_name() or user.username}"
    ws['A2'].font = Font(size=12)
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Пустая строка
    ws.append([])
    
    # Заголовки таблицы
    headers = ["Дисциплина", "Вид контроля", "Ожидаемая", "Фактическая", "В диплом"]
    ws.append(headers)
    for cell in ws[4]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Данные
    for d in disciplines:
        ws.append([
            d.title,
            d.get_assessment_type_display(),
            d.expected_grade or "—",
            d.actual_grade or "—",
            "Да" if d.for_diploma else "Нет"
        ])
    
    # Итог: средний балл
    ws.append([])
    ws.append(["Средний балл", "", "", gpa, ""])
    ws[f'A{ws.max_row}'].font = Font(bold=True)
    ws[f'D{ws.max_row}'].font = Font(bold=True, size=14, color="0000FF")
    
    # Автоширина колонок (исправление ошибки MergedCell)
    for col_idx, col in enumerate(ws.iter_cols(), start=1):
        max_length = 0
        column = openpyxl.utils.get_column_letter(col_idx)
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width if adjusted_width < 50 else 50
    
    return wb